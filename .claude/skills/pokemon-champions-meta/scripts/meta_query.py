#!/usr/bin/env python
from __future__ import annotations

import argparse
import json
import sys
from datetime import datetime
from functools import lru_cache
from pathlib import Path
from typing import Any

from meta_common import (
    details_path,
    load_json,
    maybe_repair_cn,
    norm_panel,
    normalize,
    ranking_path,
    resolve_season_rule,
)

# The sibling dex skill is the naming authority. Rather than re-read its sqlite and reimplement name
# resolution, meta reuses champdex IN-PROCESS so resolution (incl. the conservative fuzzy fallback, now
# the dex default — design: dev/update/dex/fuzzy_resolve_design.md) lives in ONE place. Absent dex
# degrades to identity (names returned unchanged). Resolved relative to the installed skills root.
_DEX_SCRIPTS = Path(__file__).resolve().parents[2] / "pokemon-champions-dex" / "scripts"
try:
    if str(_DEX_SCRIPTS) not in sys.path:
        sys.path.insert(0, str(_DEX_SCRIPTS))
    import champdex as _dex  # type: ignore
except Exception:                       # dex skill not installed alongside; degrade gracefully
    _dex = None

# panel name -> dex kind for display resolution; natures have no dex entry (left verbatim).
_KIND_BY_PANEL = {"moves": "move", "items": "item", "abilities": "ability", "partners": "pokemon"}
_DEX_TABLE = {"pokemon": "pokemon", "move": "moves", "ability": "abilities", "item": "items"}

_DEX_CONN: Any = None


def _dex_conn() -> Any:
    """One cached read-only connection to the dex DB (the DB is read-only; safe for the session)."""
    global _DEX_CONN
    if _dex is None:
        return None
    if _DEX_CONN is None:
        try:
            _DEX_CONN = _dex.conn()
        except SystemExit:              # dex DB file missing
            return None
    return _DEX_CONN

if hasattr(sys.stdout, "reconfigure"):
    sys.stdout.reconfigure(encoding="utf-8")
if hasattr(sys.stderr, "reconfigure"):
    sys.stderr.reconfigure(encoding="utf-8")


# Cached so a resident worker (worker.py / _serve.py) parses each panel file once per session
# instead of once per request. Callers treat the result as read-only. A one-shot process loads
# it once anyway, so the cache is harmless there.
@lru_cache(maxsize=None)
def load_ranking(season: str, fmt: str) -> list[dict[str, Any]]:
    return load_json(ranking_path(season, fmt), {"rows": []}).get("rows", [])


@lru_cache(maxsize=None)
def load_details(season: str, fmt: str) -> list[dict[str, Any]]:
    return load_json(details_path(season, fmt), {"rows": []}).get("rows", [])


def dex_resolve(name: str) -> tuple[set[str], dict[str, Any] | None]:
    """Resolve `name` through the dex (fuzzy by default) and return (aliases, resolution).

    `aliases` is every dex alias (any language) that shares a canonical with the resolved name, used to
    match a user query against panel rows stored in mixed languages. `resolution` is the dex fuzzy block
    — {match_type, score, distance, from} — but ONLY when the hit was a typo correction; an exact hit (or
    no dex) returns None, because there is nothing to flag. Surfacing this keeps meta honest with the dex
    rule that a fuzzy correction is never silent (dev/update/dex/fuzzy_resolve_design.md)."""
    aliases = {name}
    resolution: dict[str, Any] | None = None
    c = _dex_conn()
    if c is None:
        return aliases, resolution
    try:
        rr = _dex.resolve_rich(c, "pokemon", name)        # dex default: fuzzy-tolerant
        canon = rr.get("canonical")
        if not canon:
            return aliases, resolution
        aliases.add(canon)
        for row in c.execute("select alias from aliases where kind='pokemon' and canonical=?", (canon,)):
            if row["alias"]:
                aliases.add(str(row["alias"]))
        if rr.get("match_type") == "fuzzy":
            resolution = {"match_type": "fuzzy", "score": rr.get("score"),
                          "distance": rr.get("distance"), "from": rr.get("query_norm")}
    except Exception:
        pass
    return aliases, resolution


def dex_aliases(name: str) -> set[str]:
    """Back-compat thin wrapper: just the alias set (drops the resolution block)."""
    return dex_resolve(name)[0]


def dex_alias_norms(text: str, kinds: tuple[str, ...]) -> set[str]:
    """Normalized aliases of `text` across the given dex kinds, so a search --name in ANY language (or
    the canonical English the contract tells callers to use) matches a panel row stored only in Chinese
    + Japanese. The shipped move/item/ability panels carry no English, so without this an English query
    silently returns 0 (design promise: name lookups resolve through the dex, all three languages).
    Always includes the literal so a no-dex / unknown query still matches its own spelling."""
    out = {normalize(text)}
    c = _dex_conn()
    if c is None or not text:
        return out
    for kind in kinds:
        try:
            rr = _dex.resolve_rich(c, kind, text)        # fuzzy-tolerant, like every other dex path
            canon = rr.get("canonical")
            if not canon:
                continue
            out.add(normalize(canon))
            for row in c.execute("select alias from aliases where kind=? and canonical=?", (kind, canon)):
                if row["alias"]:
                    out.add(normalize(str(row["alias"])))
        except Exception:
            continue
    return out


def dex_canonical_type(text: str) -> str:
    """The canonical (English) type for a type query in any language (飞行 / ひこう / Flying), so the
    --type filter matches the English type stored in the data. Falls back to the literal (English types
    pass through unchanged; an unknown value stays itself and simply matches nothing)."""
    c = _dex_conn()
    if c is None or not text:
        return text
    try:
        return _dex.resolve(c, "type", text) or text
    except Exception:
        return text


@lru_cache(maxsize=None)
def _dex_display(kind: str, text: str) -> str | None:
    """Canonical Chinese display for a name of a given dex kind via the dex (fuzzy-tolerant), or None
    when the dex is unavailable or the name doesn't resolve. display_name is the dex's Chinese name."""
    c = _dex_conn()
    if c is None or not text:
        return None
    try:
        rr = _dex.resolve_rich(c, kind, text)             # dex default: fuzzy-tolerant
        canon = rr.get("canonical")
        if not canon:
            return None
        row = c.execute(f"select display_name from {_DEX_TABLE[kind]} where canonical=?", (canon,)).fetchone()
        return row["display_name"] if row and row["display_name"] else None
    except Exception:
        return None


# Garbled-CN re-decode chain (mojibake from mixed source encodings) \u2014 tried only if the dex can't
# resolve the name as-is, since a re-decoded byte string sometimes matches a real alias.
_REDECODE = (("gbk", "big5"), ("gbk", "cp950"), ("gb18030", "big5"), ("gb18030", "cp950"),
             ("big5", "gbk"), ("cp950", "gbk"), ("big5", "gb18030"), ("cp950", "gb18030"))


def repair_display_name(name: Any, kind: str | None = None) -> str:
    """Best Chinese display for a scraped name. Asks the dex (the naming authority) for the canonical
    display of the given kind, fuzzy-tolerant; `kind` is None for entries with no dex entity (natures),
    which are returned as-is. Falls back to the source text when the dex can't resolve it."""
    text = maybe_repair_cn(str(name or ""))
    if not text or not kind:
        return text
    disp = _dex_display(kind, text)
    if disp:
        return disp
    for source, target in _REDECODE:
        try:
            candidate = text.encode(source).decode(target)
        except Exception:
            continue
        disp = _dex_display(kind, candidate)
        if disp:
            return disp
    return text


def _flag_resolution(detail: dict[str, Any], name: str,
                     resolution: dict[str, Any] | None) -> dict[str, Any]:
    """Attach the typo-correction metadata to a COPY of the matched panel (the panel is shared lru-cache
    state — never mutate it). No-op when the query resolved exactly. Surfaces `query` (raw input),
    `resolved_name` (English canonical of the matched panel), and `name_resolution` (the dex fuzzy block)
    so a caller can distinguish an exact lookup from an auto-corrected typo."""
    if not resolution:
        return detail
    resolved = detail.get("pokemon_en") or detail.get("slug") or ""
    return {**detail, "query": name, "resolved_name": resolved,
            "name_resolution": {**resolution, "resolved": resolved}}


def _resolve_one(name: str, ranking: list[dict[str, Any]],
                 details_by_slug: dict[str, dict[str, Any]]) -> dict[str, Any] | None:
    """Resolve one name against already-loaded ranking + details (no file I/O here). A fuzzy (typo) hit
    carries a non-silent `name_resolution` block on the returned copy; an exact hit returns the panel as-is."""
    aliases, resolution = dex_resolve(name)
    needles = {normalize(x) for x in aliases}
    needles.discard("")  # never match on empty fields
    for row in ranking:
        values = [
            row.get("slug", ""),
            row.get("pokemon", ""),
            row.get("alt_name", ""),
            row.get("pokemon_en", ""),
            row.get("pokemon_ja", ""),
        ]
        if any(normalize(v) in needles for v in values):
            return _flag_resolution(details_by_slug.get(row.get("slug")) or row, name, resolution)
    for row in details_by_slug.values():
        values = [row.get("slug", ""), row.get("pokemon", ""), row.get("pokemon_en", ""), row.get("pokemon_ja", "")]
        if any(normalize(v) in needles for v in values):
            return _flag_resolution(row, name, resolution)
    return None


def resolve_many(names: list[str], season: str, fmt: str) -> list[dict[str, Any] | None]:
    """Resolve several names against the season/format panels, loading the data files ONCE.
    Returns a list aligned to `names` (None where a name doesn't resolve)."""
    ranking = load_ranking(season, fmt)
    details_by_slug = {row.get("slug"): row for row in load_details(season, fmt)}
    return [_resolve_one(n, ranking, details_by_slug) for n in names]


def resolve_pokemon(name: str, season: str, fmt: str) -> dict[str, Any] | None:
    return resolve_many([name], season, fmt)[0]


# --- Canonical I/O contract (dev/contracts/conventions.md) ----------------------------------------
# Applied to every JSON CLI result via emit(): pokemon rows expose the join key `name` (English) plus
# localized `name_zh`/`name_ja` alongside the existing pokemon_*/slug (kept for back-compat), and move
# type/category are Title-cased to match dex/calc. Panel entries own `name` (the move/ability/item), so
# pokemon-identity keys are added ONLY to pure pokemon rows (slug present, no panel `key`) — never to
# the flattened search rows, where `name` is the matched entry. Data files are left untouched.
_ID_CANON = (("pokemon_en", "name"), ("pokemon", "name_zh"), ("pokemon_ja", "name_ja"))


def canonicalize(obj: Any) -> Any:
    if isinstance(obj, list):
        return [canonicalize(x) for x in obj]
    if isinstance(obj, dict):
        out = {k: (canonicalize(v) if isinstance(v, (dict, list)) else v) for k, v in obj.items()}
        if "slug" in out and "key" not in out:           # a pure pokemon row, not a panel/search row
            for src, dst in _ID_CANON:
                if out.get(src) and dst not in out:
                    out[dst] = out[src]
        for f in ("type", "category"):                   # match dex/calc Title-case
            if isinstance(out.get(f), str) and out[f]:
                out[f] = out[f].capitalize()
        return out
    return obj


def emit(data: Any, output: str) -> None:
    if output == "json":
        print(json.dumps(canonicalize(data), ensure_ascii=False, indent=2))
    else:
        print(data if isinstance(data, str) else to_markdown(data))


def ranking_md(rows: list[dict[str, Any]]) -> str:
    lines = ["| rank | pokemon | slug | en |", "|---:|---|---|---|"]
    for row in rows:
        lines.append(f"| {row.get('rank','')} | {row.get('pokemon','')} | {row.get('slug','')} | {row.get('pokemon_en','')} |")
    return "\n".join(lines)


def detail_md(detail: dict[str, Any], panel: str | None = None) -> str:
    lines = [
        f"# {detail.get('pokemon') or detail.get('slug')} ({detail.get('format')}, {detail.get('season')}/{detail.get('rule', '')})",
        "",
        f"- rank: {detail.get('rank','')}",
        f"- slug: {detail.get('slug','')}",
        f"- en: {detail.get('pokemon_en','')}",
    ]
    nr = detail.get("name_resolution")
    if nr:  # non-silent typo correction: tell the reader the query was auto-resolved
        lines.append(f"- resolved: `{detail.get('query','')}` → {detail.get('resolved_name','')} "
                     f"(fuzzy, distance {nr.get('distance')})")
    lines.append("")
    panels = detail.get("panels", {})
    selected = [norm_panel(panel)] if panel else ["moves", "items", "abilities", "natures", "partners", "spreads"]
    for key in selected:
        entries = panels.get(key, [])
        lines.extend([f"## {key}", "", "| rank | name | ja/key | usage | extra |", "|---:|---|---|---:|---|"])
        for entry in entries:
            extra = []
            for k in ("type", "category", "power", "accuracy"):
                if entry.get(k) not in ("", None):
                    extra.append(f"{k}={entry.get(k)}")
            lines.append(
                f"| {entry.get('rank','')} | {entry.get('name','')} | {entry.get('name_ja') or entry.get('key','')} | "
                f"{percent_text(entry.get('percentage'))} | {', '.join(extra)} |"
            )
        if not entries:
            lines.append("|  |  |  |  |  |")
        lines.append("")
    return "\n".join(lines)


def search_md(rows: list[dict[str, Any]]) -> str:
    lines = ["| pokemon_rank | pokemon | slug | format | panel | entry_rank | name | usage |", "|---:|---|---|---|---|---:|---|---:|"]
    for row in rows:
        lines.append(
            f"| {row.get('pokemon_rank','')} | {row.get('pokemon','')} | {row.get('slug','')} | {row.get('format','')} | "
            f"{row.get('panel','')} | {row.get('entry_rank','')} | {row.get('name','')} | {percent_text(row.get('percentage'))} |"
        )
    return "\n".join(lines)


def to_markdown(data: Any) -> str:
    if isinstance(data, list):
        return search_md(data)
    if isinstance(data, dict):
        return detail_md(data)
    return str(data)


def command_ranking(args: argparse.Namespace) -> None:
    args.season, args.rule = resolve_season_rule(args.season, args.rule)
    rows = load_ranking(args.season, args.format)[: args.limit]
    emit({"season": args.season, "rule": args.rule, "format": args.format, "rows": rows} if args.output == "json" else ranking_md(rows), args.output)


def _panel_filtered(detail: dict[str, Any], panel: str | None) -> dict[str, Any]:
    if not panel:
        return detail
    return {**detail, "panels": {norm_panel(panel): detail.get("panels", {}).get(norm_panel(panel), [])}}


def command_detail(args: argparse.Namespace) -> None:
    args.season, args.rule = resolve_season_rule(args.season, args.rule)
    names = args.pokemon if isinstance(args.pokemon, list) else [args.pokemon]
    resolved = resolve_many(names, args.season, args.format)
    if len(names) == 1:
        # Single name keeps the original scalar shape (object for json, md report).
        detail = resolved[0]
        if not detail:
            msg = f"not found: {names[0]} in {args.season} {args.format}"
            if args.output == "json":
                # JSON mode must emit JSON, not a bare SystemExit string (a caller parses stdout).
                emit({"ok": False, "query": names[0], "error": {"code": "not_found", "message": msg}}, "json")
                raise SystemExit(1)
            raise SystemExit(msg)
        emit(_panel_filtered(detail, args.panel) if args.output == "json" else detail_md(detail, args.panel),
             args.output)
        return
    # Batch: resolve every name in ONE invocation (data files loaded once). Aligned to `names`;
    # an unresolved name becomes the uniform error object rather than aborting the whole batch.
    if args.output == "json":
        # Batch errors carry `index` to align with request order (conventions.md error shape; audit
        # 2026-06-24 — only `query` was present before).
        out = [_panel_filtered(d, args.panel) if d
               else {"ok": False, "index": i, "query": n,
                     "error": {"code": "not_found", "message": f"not found: {n} in {args.season} {args.format}"}}
               for i, (n, d) in enumerate(zip(names, resolved))]
        emit(out, "json")
    else:
        blocks = [detail_md(d, args.panel) if d else f"# {n}\n- not found" for n, d in zip(names, resolved)]
        print("\n\n---\n\n".join(blocks))


def iter_panel_rows(season: str, fmt: str, panel: str | None) -> list[dict[str, Any]]:
    rows: list[dict[str, Any]] = []
    panels = [norm_panel(panel)] if panel else ["moves", "items", "abilities", "natures", "partners", "spreads"]
    for detail in load_details(season, fmt):
        for key in panels:
            for entry in detail.get("panels", {}).get(key, []) or []:
                row = {
                    "pokemon_rank": detail.get("rank"),
                    "pokemon": detail.get("pokemon"),
                    "slug": detail.get("slug"),
                    "format": fmt,
                    "season": season,
                    "rule": detail.get("rule", ""),
                    "panel": key,
                    "entry_rank": entry.get("rank"),
                    "name": entry.get("name"),
                    "name_ja": entry.get("name_ja"),
                    "percentage": entry.get("percentage"),
                }
                row.update({k: v for k, v in entry.items() if k not in row})
                rows.append(row)
    return rows


def percent_text(value: Any) -> str:
    if value in ("", None):
        return ""
    try:
        return f"{float(value):g}%"
    except Exception:
        return str(value)


def ranked_prefix(entry: dict[str, Any]) -> str:
    rank = entry.get("rank")
    return f"{rank}. " if rank not in ("", None) else ""


NATURE_EFFECTS = {
    "さみしがり": ("A", "B"),
    "ゆうかん": ("A", "S"),
    "いじっぱり": ("A", "C"),
    "やんちゃ": ("A", "D"),
    "ずぶとい": ("B", "A"),
    "のんき": ("B", "S"),
    "わんぱく": ("B", "C"),
    "のうてんき": ("B", "D"),
    "おくびょう": ("S", "A"),
    "せっかち": ("S", "B"),
    "ようき": ("S", "C"),
    "むじゃき": ("S", "D"),
    "ひかえめ": ("C", "A"),
    "おっとり": ("C", "B"),
    "れいせい": ("C", "S"),
    "うっかりや": ("C", "D"),
    "おだやか": ("D", "A"),
    "おとなしい": ("D", "B"),
    "なまいき": ("D", "S"),
    "しんちょう": ("D", "C"),
    "怕寂寞": ("A", "B"),
    "勇敢": ("A", "S"),
    "固执": ("A", "C"),
    "顽皮": ("A", "D"),
    "大胆": ("B", "A"),
    "悠闲": ("B", "S"),
    "淘气": ("B", "C"),
    "乐天": ("B", "D"),
    "胆小": ("S", "A"),
    "急躁": ("S", "B"),
    "爽朗": ("S", "C"),
    "天真": ("S", "D"),
    "内敛": ("C", "A"),
    "慢吞吞": ("C", "B"),
    "冷静": ("C", "S"),
    "马虎": ("C", "D"),
    "温和": ("D", "A"),
    "温顺": ("D", "B"),
    "自大": ("D", "S"),
    "慎重": ("D", "C"),
}


def nature_effect_text(entry: dict[str, Any], name: str) -> str:
    effect = NATURE_EFFECTS.get(str(entry.get("key") or "")) or NATURE_EFFECTS.get(str(entry.get("name_ja") or "")) or NATURE_EFFECTS.get(name)
    if not effect:
        return ""
    up, down = effect
    return f" +{up} -{down}"


def format_named_entry(entry: dict[str, Any], panel: str) -> str:
    name = repair_display_name(entry.get("name") or entry.get("name_ja") or entry.get("key") or "",
                              _KIND_BY_PANEL.get(panel))
    if panel == "partners":
        return f"{ranked_prefix(entry)}{name}"
    usage = percent_text(entry.get("percentage"))
    if panel == "natures":
        effect = nature_effect_text(entry, name)
        suffix = f"{effect} ({usage})" if usage else effect
        return f"{ranked_prefix(entry)}{name}{suffix}"
    suffix = f" ({usage})" if usage else ""
    return f"{ranked_prefix(entry)}{name}{suffix}"


def format_spread_entry(entry: dict[str, Any]) -> str:
    usage = percent_text(entry.get("percentage"))
    values = [
        f"H{entry.get('hp', '')}",
        f"A{entry.get('atk', '')}",
        f"B{entry.get('def', '')}",
        f"C{entry.get('spa', '')}",
        f"D{entry.get('spd', '')}",
        f"S{entry.get('spe', '')}",
    ]
    suffix = f" ({usage})" if usage else ""
    return f"{ranked_prefix(entry)}{' / '.join(values)}{suffix}"


def format_panel(entries: list[dict[str, Any]], panel: str) -> str:
    lines: list[str] = []
    for entry in entries or []:
        if panel == "spreads":
            lines.append(format_spread_entry(entry))
        else:
            lines.append(format_named_entry(entry, panel))
    return "\n".join(lines)


def export_human_rows(season: str, fmt: str) -> list[dict[str, str]]:
    rows: list[dict[str, str]] = []
    ranking_by_slug = {row.get("slug"): row for row in load_ranking(season, fmt)}
    details = sorted(load_details(season, fmt), key=lambda row: row.get("rank") or 999999)
    for detail in details:
        ranking = ranking_by_slug.get(detail.get("slug"), {})
        rank = detail.get("rank") or ranking.get("rank")
        name = repair_display_name(detail.get("pokemon") or ranking.get("pokemon") or detail.get("slug") or "",
                                   "pokemon")
        en = detail.get("pokemon_en") or ranking.get("pokemon_en") or ""
        panels = detail.get("panels", {})
        rows.append(
            {
                "排名": rank,
                "中文名": name,
                "英文名": en,
                "招式": format_panel(panels.get("moves", []), "moves"),
                "道具": format_panel(panels.get("items", []), "items"),
                "特性": format_panel(panels.get("abilities", []), "abilities"),
                "性格": format_panel(panels.get("natures", []), "natures"),
                "队友": format_panel(panels.get("partners", []), "partners"),
                "努力值": format_panel(panels.get("spreads", []), "spreads"),
            }
        )
    return rows


def default_excel_path(season: str, output_dir: Path | None = None) -> Path:
    export_date = datetime.now().strftime("%Y%m%d")
    return (output_dir or Path.cwd()) / f"{season}_{export_date}.xlsx"


def project_root() -> Path:
    return Path(__file__).resolve().parents[4]


# Data-sheet columns. 队友 (partners) is deliberately the last data column so the
# lookup sheet's name search never collides with partner names.
DATA_COLUMNS = ["排名", "中文名", "英文名", "招式", "道具", "特性", "性格", "队友", "努力值"]
DATA_WIDTHS = {"排名": 8, "中文名": 18, "英文名": 20, "招式": 26, "道具": 24,
               "特性": 24, "性格": 28, "队友": 24, "努力值": 42}
FONT = "汉仪旗黑-55S"
INK = "1F2937"
FMT_SHEET = {"single": "单打", "double": "双打"}


def _style_data_sheet(ws, rows: list[dict[str, str]]) -> None:
    from openpyxl.styles import Alignment, Border, Font, PatternFill, Side
    from openpyxl.utils import get_column_letter

    ws.sheet_view.showGridLines = False
    ws.append(DATA_COLUMNS)
    for row in rows:
        ws.append([row.get(c, "") for c in DATA_COLUMNS])
    ws.freeze_panes = "A2"
    ws.auto_filter.ref = ws.dimensions
    for idx, column in enumerate(DATA_COLUMNS, 1):
        ws.column_dimensions[get_column_letter(idx)].width = DATA_WIDTHS[column]
    header_fill = PatternFill("solid", fgColor="D9EAF7")
    band_fill = PatternFill("solid", fgColor="F7FBFD")
    border = Border(bottom=Side(style="thin", color="D6E3EA"))
    for row_idx, row in enumerate(ws.iter_rows(), 1):
        for cell in row:
            cell.font = Font(name=FONT, size=10, bold=(row_idx == 1), color=INK)
            cell.border = border
            cell.alignment = Alignment(horizontal="center", vertical="center", wrap_text=True)
            if row_idx == 1:
                cell.fill = header_fill
            elif row_idx % 2 == 0:
                cell.fill = band_fill
    ws.row_dimensions[1].height = 22


def _load_report(report_dir: Path | None, season: str, fmt: str) -> dict[str, Any] | None:
    if not report_dir:
        return None
    path = report_dir / f"report_{season}_{fmt}.json"
    if not path.exists():
        return None
    try:
        return json.loads(path.read_text(encoding="utf-8"))
    except Exception:
        return None


def _write_report_table(ws, top: int, left: int, title: str, headers: list[str],
                        rows: list[list[Any]]) -> int:
    """Render one titled table starting at (top,left); return the next free row."""
    from openpyxl.styles import Alignment, Font, PatternFill
    from openpyxl.utils import get_column_letter

    tcell = ws.cell(row=top, column=left, value=f"{title}（{len(rows)}）")
    tcell.font = Font(name=FONT, size=11, bold=True, color="0B5394")
    top += 1
    for j, h in enumerate(headers):
        c = ws.cell(row=top, column=left + j, value=h)
        c.font = Font(name=FONT, size=10, bold=True, color="FFFFFF")
        c.fill = PatternFill("solid", fgColor="6FA8DC")
        c.alignment = Alignment(horizontal="center", vertical="center", wrap_text=True)
    top += 1
    if not rows:
        ws.cell(row=top, column=left, value="（无）").font = Font(name=FONT, size=10, italic=True, color="888888")
        return top + 2
    for r, rowvals in enumerate(rows):
        band = (r % 2 == 1)
        for j, v in enumerate(rowvals):
            c = ws.cell(row=top + r, column=left + j, value=v)
            c.font = Font(name=FONT, size=10, color=INK)
            c.alignment = Alignment(horizontal="center", vertical="center")
            if band:
                c.fill = PatternFill("solid", fgColor="F2F7FB")
    return top + len(rows) + 2


def _write_report_block(ws, top: int, left: int, title: str, rep: dict[str, Any]) -> int:
    from openpyxl.styles import Alignment, Font, PatternFill
    from openpyxl.utils import get_column_letter

    ws.merge_cells(start_row=top, start_column=left, end_row=top, end_column=left + 6)
    head = ws.cell(row=top, column=left, value=title)
    head.font = Font(name=FONT, size=13, bold=True, color="FFFFFF")
    head.fill = PatternFill("solid", fgColor="0B5394")
    head.alignment = Alignment(horizontal="left", vertical="center")
    ws.cell(row=top + 1, column=left,
            value=f"对比基线 {rep.get('baseline', {}).get('kind', '?')}　当前 {str(rep.get('current', {}).get('details_updated_at', '?'))[:10]}").font = Font(name=FONT, size=9, italic=True, color="666666")
    top += 3
    rm = [[r["name"], r["old"], r["new"], ("↑" if r["delta"] > 0 else "↓") + str(abs(r["delta"])), r["tier"]] for r in rep.get("rank_moves", [])]
    top = _write_report_table(ws, top, left, "名次显著变动", ["宝可梦", "旧名次", "新名次", "Δ", "档"], rm)
    ne = [[r["name"], r["rank"], r["tier"]] for r in rep.get("new_entries", [])]
    top = _write_report_table(ws, top, left, "新进榜", ["宝可梦", "新名次", "档"], ne)
    dp = [[r["name"], r["old"], r["tier"]] for r in rep.get("dropped", [])]
    top = _write_report_table(ws, top, left, "跌出榜", ["宝可梦", "旧名次", "档"], dp)
    # 面板 + 项目 merged into one column ("moves/仆刀") to keep the table narrow.
    cc = [[f"{c['name']}({c['rank']})", f"{c['panel']}/{c['item']}", c["old_pct"], c["new_pct"],
           f"{c['delta']:+.1f}", c["tier"], "是" if c["rank_triggered"] else ""] for c in rep.get("config_changes", [])]
    top = _write_report_table(ws, top, left, "配置显著变动", ["宝可梦(名次)", "面板/项目", "旧%", "新%", "Δpp", "档", "触发"], cc)
    return top


def _build_report_sheet(wb, season: str, single_rows: list[dict[str, str]],
                        double_rows: list[dict[str, str]], reports: dict[str, Any]) -> None:
    from openpyxl.styles import Alignment, Border, Font, PatternFill, Side
    from openpyxl.worksheet.datavalidation import DataValidation

    ws = wb.create_sheet("更新报告")
    ws.sheet_view.showGridLines = False
    thin = Border(*(Side(style="thin", color="C9D6DF"),) * 4)
    # column bands: single report A-G (1-7), gap H, double report I-O (9-15)
    for col, w in (("A", 22), ("B", 22), ("C", 8), ("D", 8), ("E", 8), ("F", 6), ("G", 9),
                   ("H", 3), ("I", 22), ("J", 22), ("K", 8), ("L", 8), ("M", 8), ("N", 6), ("O", 9)):
        ws.column_dimensions[col].width = w

    # --- interactive lookup: 单打 spans B:G, 双打 spans I:O (same bands as reports) ---
    ws.merge_cells("A1:G1")
    title = ws.cell(row=1, column=1, value="宝可梦检索（单 / 双 对照）")
    title.font = Font(name=FONT, size=14, bold=True, color="0B5394")
    ws.cell(row=2, column=1, value="输入中文名 →").font = Font(name=FONT, size=11, bold=True, color=INK)
    ws.merge_cells("B2:G2")
    inp = ws.cell(row=2, column=2, value=(single_rows[0]["中文名"] if single_rows else ""))
    inp.font = Font(name=FONT, size=12, bold=True, color="990000")
    inp.fill = PatternFill("solid", fgColor="FFF2CC")
    inp.border = thin
    inp.alignment = Alignment(horizontal="center", vertical="center")

    # dropdown source: union of names, hidden far-right in column Q
    names = list(dict.fromkeys([r["中文名"] for r in single_rows] + [r["中文名"] for r in double_rows]))
    for i, nm in enumerate(names, start=1):
        ws.cell(row=i, column=17, value=nm)  # col Q
    ws.column_dimensions["Q"].hidden = True
    dv = DataValidation(type="list", formula1=f"'更新报告'!$Q$1:$Q${max(len(names),1)}", allow_blank=True)
    ws.add_data_validation(dv)
    dv.add(inp)

    # header row
    ws.cell(row=4, column=1, value="项目").font = Font(name=FONT, size=11, bold=True, color="FFFFFF")
    ws.cell(row=4, column=1).fill = PatternFill("solid", fgColor="6FA8DC")
    ws.cell(row=4, column=1).alignment = Alignment(horizontal="center", vertical="center")
    for band, lbl in ((2, "单打（B:G）"), (9, "双打（I:O）")):
        ws.merge_cells(start_row=4, start_column=band, end_row=4, end_column=band + 5)
        h = ws.cell(row=4, column=band, value=lbl.split("（")[0])
        h.font = Font(name=FONT, size=11, bold=True, color="FFFFFF")
        h.fill = PatternFill("solid", fgColor="6FA8DC")
        h.alignment = Alignment(horizontal="center", vertical="center")

    # 项目 rows; 队友 deliberately last. MATCH on 中文名 (col B) of each data sheet.
    layout = [("排名", "A"), ("英文名", "C"), ("招式", "D"), ("道具", "E"),
              ("特性", "F"), ("性格", "G"), ("努力值", "I"), ("队友", "H")]
    for r, (label, col) in enumerate(layout, start=5):
        lc = ws.cell(row=r, column=1, value=label)
        lc.font = Font(name=FONT, size=10, bold=True, color=INK)
        lc.fill = PatternFill("solid", fgColor="EAF1F8")
        lc.alignment = Alignment(horizontal="center", vertical="top")
        lc.border = thin
        for band, sheet in ((2, "单打"), (9, "双打")):
            ws.merge_cells(start_row=r, start_column=band, end_row=r, end_column=band + 5)
            f = (f"=IFERROR(INDEX('{sheet}'!{col}:{col},MATCH($B$2,'{sheet}'!$B:$B,0)),\"—\")")
            c = ws.cell(row=r, column=band, value=f)
            c.font = Font(name=FONT, size=10, color=INK)
            c.alignment = Alignment(horizontal="left", vertical="top", wrap_text=True)
        ws.row_dimensions[r].height = 80 if label in ("招式", "道具", "性格", "努力值", "队友") else 18

    # --- update reports: single (A-G) and double (I-O) side by side ---
    top = 5 + len(layout) + 2
    rep_s, rep_d = reports.get("single"), reports.get("double")
    if rep_s:
        _write_report_block(ws, top, 1, "单打更新报告", rep_s)
    if rep_d:
        _write_report_block(ws, top, 9, "双打更新报告", rep_d)


def command_export_excel(args: argparse.Namespace) -> None:
    try:
        from openpyxl import Workbook
    except Exception as exc:
        raise SystemExit("export-excel requires openpyxl to be installed") from exc

    args.season, args.rule = resolve_season_rule(args.season, args.rule)
    output_dir = Path(args.output_dir) if args.output_dir else Path.cwd()
    output_file = Path(args.output_file) if args.output_file else default_excel_path(args.season, output_dir)
    output_file.parent.mkdir(parents=True, exist_ok=True)
    # The update report ships inside the skill's data dir (report_<season>_<fmt>.json),
    # so it is embedded by default — no --report-dir needed.
    report_dir = Path(args.report_dir) if args.report_dir else details_path(args.season, "single").parent

    rows_by_fmt = {fmt: export_human_rows(args.season, fmt) for fmt in ("single", "double")}
    reports = {fmt: _load_report(report_dir, args.season, fmt) for fmt in ("single", "double")}

    wb = Workbook()
    wb.remove(wb.active)
    for fmt in ("single", "double"):
        ws = wb.create_sheet(FMT_SHEET[fmt])
        _style_data_sheet(ws, rows_by_fmt[fmt])
    _build_report_sheet(wb, args.season, rows_by_fmt["single"], rows_by_fmt["double"], reports)
    wb.save(output_file)

    result = {
        "output_file": str(output_file),
        "sheets": [FMT_SHEET["single"], FMT_SHEET["double"], "更新报告"],
        "single_rows": len(rows_by_fmt["single"]),
        "double_rows": len(rows_by_fmt["double"]),
        "reports_embedded": [f for f, r in reports.items() if r],
    }
    print(json.dumps(result, ensure_ascii=False, indent=2))


def command_search(args: argparse.Namespace) -> None:
    args.season, args.rule = resolve_season_rule(args.season, args.rule)
    formats = ["single", "double"] if args.format == "both" else [args.format]
    results: list[dict[str, Any]] = []
    # Resolve --name through the dex so an English/Chinese/JP query matches a panel row stored in another
    # language (the move/item/ability panels carry no English). The kinds searched come from --panel; a
    # bare search (no panel) covers every entity panel. Each needle expands to its cross-language alias
    # set and matches if ANY alias is present — done once per query, not per row.
    if args.panel:
        p = norm_panel(args.panel)
        name_kinds = (_KIND_BY_PANEL[p],) if p in _KIND_BY_PANEL else ()
    else:
        name_kinds = ("move", "item", "ability", "pokemon")
    name_alias_sets = [dex_alias_norms(x, name_kinds) for x in args.name]
    wanted_type = dex_canonical_type(args.type) if args.type else None
    for fmt in formats:
        for row in iter_panel_rows(args.season, fmt, args.panel):
            hay = normalize(" ".join(str(row.get(k, "")) for k in ("name", "name_ja", "key", "slug", "pokemon")))
            if name_alias_sets and not all(any(a in hay for a in aset) for aset in name_alias_sets):
                continue
            if wanted_type and normalize(row.get("type", "")) != normalize(wanted_type):
                continue
            if args.category and normalize(row.get("category", "")) != normalize(args.category):
                continue
            if args.min_usage is not None:
                try:
                    if float(row.get("percentage") or 0) < args.min_usage:
                        continue
                except Exception:
                    continue
            results.append(row)
    results.sort(key=lambda r: (r.get("format", ""), r.get("pokemon_rank") or 9999, r.get("panel", ""), r.get("entry_rank") or 9999))
    results = results[: args.limit]
    # JSON output is a bare list of matched panel rows (consistent with `detail`); the season/rule are
    # query inputs the caller already knows, and each row carries its own format/panel/identity.
    emit(results if args.output == "json" else search_md(results), args.output)


def command_compare(args: argparse.Namespace) -> None:
    args.season, args.rule = resolve_season_rule(args.season, args.rule)
    rows = []
    resolution = None        # captured from whichever format resolves (same query for both)
    resolved_name = None
    for fmt in ("single", "double"):
        detail = resolve_pokemon(args.pokemon, args.season, fmt)
        if not detail:
            rows.append({"format": fmt, "found": False})
            continue
        if detail.get("name_resolution") and resolution is None:
            resolution = detail["name_resolution"]
            resolved_name = detail.get("resolved_name")
        panels = detail.get("panels", {})
        rows.append(
            {
                "format": fmt,
                "found": True,
                "rank": detail.get("rank"),
                "pokemon": detail.get("pokemon"),
                # Thread the English/JP identity so canonicalize emits `name` — the cross-skill JOIN key
                # (every skill must carry it). Without pokemon_en the row had only pokemon/slug, so `name`
                # vanished and a caller couldn't feed compare output to calc/dex (audit 2026-06-28).
                "pokemon_en": detail.get("pokemon_en"),
                "pokemon_ja": detail.get("pokemon_ja"),
                "slug": detail.get("slug"),
                "top_moves": [x.get("name") for x in panels.get("moves", [])[:5]],
                "top_items": [x.get("name") for x in panels.get("items", [])[:5]],
                "top_abilities": [x.get("name") for x in panels.get("abilities", [])[:3]],
                "top_partners": [x.get("name") for x in panels.get("partners", [])[:5]],
            }
        )
    if args.output == "json":
        # Key the per-format results by format ("single"/"double") so callers can address them
        # directly (compared["single"]); season/rule/pokemon stay as sibling metadata.
        out = {"season": args.season, "rule": args.rule, "pokemon": args.pokemon}
        if resolution:  # non-silent typo correction
            out["resolved_name"] = resolved_name
            out["name_resolution"] = resolution
        for row in rows:
            out[row["format"]] = row
        emit(out, args.output)
        return
    lines = [f"# Compare {args.pokemon}", ""]
    if resolution:
        lines.append(f"- resolved: `{args.pokemon}` → {resolved_name} (fuzzy, distance {resolution.get('distance')})")
        lines.append("")
    for row in rows:
        if not row.get("found"):
            lines.append(f"## {row['format']}\nnot found\n")
            continue
        lines.append(f"## {row['format']}")
        lines.append(f"- rank: {row.get('rank')}")
        lines.append(f"- pokemon: {row.get('pokemon')} ({row.get('slug')})")
        lines.append(f"- moves: {'; '.join(row.get('top_moves', []))}")
        lines.append(f"- items: {'; '.join(row.get('top_items', []))}")
        lines.append(f"- abilities: {'; '.join(row.get('top_abilities', []))}")
        lines.append(f"- partners: {'; '.join(row.get('top_partners', []))}")
        lines.append("")
    print("\n".join(lines))


# Machine-readable I/O contract (dev/contracts/conventions.md), emitted by `schema`. Pinned by the
# meta contract test so it can't drift from what the CLI actually returns.
META_SCHEMA = {
    "skill": "pokemon-champions-meta",
    "contract": "dev/contracts/conventions.md",
    "identity": {"name": "English join key", "name_zh": "中文", "name_ja": "日本語", "slug": "lowercase"},
    "stat_keys": ["hp", "atk", "def", "spa", "spd", "spe"],
    "casing": "move type/category are Title-case",
    "error_shape": {"ok": False, "query": "<input>", "error": {"code": "not_found", "message": "<str>"}},
    "commands": {
        "ranking": "--format single|double [--season --rule --limit] -> {season,rule,format,rows:[pokemon row]}",
        "detail": "--format --pokemon <name ...> [--panel] -> detail(1) | list(>1; misses -> error_shape)",
        "search": "[--panel --name --type --category --min-usage --limit] -> [flattened panel rows]",
        "compare": "--pokemon <name> -> {single:{found,...},double:{found,...}}  (found:false = not ranked there)",
        "export-excel": "[--output-file/-dir/-report-dir] writes the data + factual-change workbook (单打/双打/更新报告 sheets)",
        "schema": "this contract",
    },
    "shapes": {
        "ranking_row": "{rank, name, name_zh, name_ja, slug, pokemon, pokemon_en, pokemon_ja}",
        "detail": "{rank, name, name_zh, name_ja, slug, panels:{moves,items,abilities,natures,partners,spreads}}",
        "move_entry": "{rank, name, name_ja, key, percentage:float, type:Title, category:Title, power:int|null}",
        "spread_entry": "{rank, hp,atk,def,spa,spd,spe:int, percentage:float}",
        "name_resolution": "on a TYPO'd query only, detail/compare add {query, resolved_name, name_resolution:{match_type:'fuzzy', score, distance, from}}; an exact query omits all three",
    },
}


def command_schema(args: argparse.Namespace) -> None:
    print(json.dumps(META_SCHEMA, ensure_ascii=False, indent=2))


def main() -> int:
    parser = argparse.ArgumentParser()
    sub = parser.add_subparsers(dest="command", required=True)
    schema_p = sub.add_parser("schema")
    schema_p.add_argument("--output", choices=["md", "json"], default="json")  # ignored; uniform CLI
    schema_p.set_defaults(func=command_schema)

    ranking = sub.add_parser("ranking")
    ranking.add_argument("--format", choices=["single", "double"], required=True)
    ranking.add_argument("--season")
    ranking.add_argument("--rule")
    ranking.add_argument("--limit", type=int, default=20)
    ranking.add_argument("--output", choices=["md", "json"], default="md")
    ranking.set_defaults(func=command_ranking)

    detail = sub.add_parser("detail")
    detail.add_argument("--format", choices=["single", "double"], required=True)
    detail.add_argument("--season")
    detail.add_argument("--rule")
    detail.add_argument("--pokemon", required=True, nargs="+",
                        help="one or more names; with >1 name the json output is a list (data files "
                             "loaded once) so callers can fetch many panels in a single process")
    detail.add_argument("--panel")
    detail.add_argument("--output", choices=["md", "json"], default="md")
    detail.set_defaults(func=command_detail)

    search = sub.add_parser("search")
    search.add_argument("--format", choices=["single", "double", "both"], default="both")
    search.add_argument("--season")
    search.add_argument("--rule")
    search.add_argument("--panel")
    search.add_argument("--name", action="append", default=[])
    search.add_argument("--type")
    search.add_argument("--category")
    search.add_argument("--min-usage", type=float)
    search.add_argument("--limit", type=int, default=50)
    search.add_argument("--output", choices=["md", "json"], default="md")
    search.set_defaults(func=command_search)

    compare = sub.add_parser("compare")
    compare.add_argument("--season")
    compare.add_argument("--rule")
    compare.add_argument("--pokemon", required=True)
    compare.add_argument("--output", choices=["md", "json"], default="md")
    compare.set_defaults(func=command_compare)

    export_excel = sub.add_parser("export-excel")
    export_excel.add_argument("--season")
    export_excel.add_argument("--rule")
    export_excel.add_argument("--output-file", help="Explicit .xlsx path for the combined single+double workbook.")
    export_excel.add_argument("--output-dir", help="Directory for the default export name; defaults to the current working directory.")
    export_excel.add_argument("--report-dir", help="Directory holding report_<season>_<fmt>.json to embed in the 更新报告 sheet. Defaults to the skill's own data dir.")
    export_excel.set_defaults(func=command_export_excel)

    args = parser.parse_args()
    args.func(args)
    return 0


if __name__ == "__main__":
    try:
        raise SystemExit(main())
    except BrokenPipeError:
        sys.exit(0)
